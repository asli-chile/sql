"""
Módulo para exportar datos normalizados a Excel.
"""
import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Exportador de datos a formato Excel."""
    
    def __init__(self):
        """Inicializa el exportador Excel."""
        pass
    
    def export_single(self, data: Dict[str, any], output_path: str) -> str:
        """
        Exporta un único registro a Excel.
        Si hay múltiples naves, exporta todas.
        
        Args:
            data: Diccionario con datos normalizados.
            output_path: Ruta donde guardar el archivo Excel.
            
        Returns:
            Ruta del archivo generado.
        """
        output_path_obj = Path(output_path)
        output_path_obj.parent.mkdir(parents=True, exist_ok=True)
        
        # Si hay múltiples naves, usar export_multiple
        if data.get('multiple_naves') and data.get('naves_normalizadas'):
            return self.export_multiple(data['naves_normalizadas'], output_path, data)
        
        # Una sola nave - exportación normal con campos requeridos
        # Crear DataFrame con los datos
        df_data = {
            'Campo': [],
            'Valor': []
        }
        
        # Campos siempre requeridos (en orden específico)
        fields_mapping = {
            'Naviera': data.get('naviera'),
            'Nave': data.get('nave_normalizada') or data.get('nave_original'),
            'POL': data.get('pol'),
            'POD': data.get('pod'),
            'ETD': data.get('etd') or data.get('fecha_salida_normalizada'),
            'ETA': data.get('eta') or data.get('fecha_llegada_normalizada'),
            'Fecha': data.get('fecha_normalizada'),
            'Semana': data.get('semana_normalizada'),
            'Número de Contenedor': data.get('numero_contenedor'),
            'Número de Booking': data.get('numero_booking'),
        }
        
        # Agregar campos adicionales si existen
        if 'numero_viaje' in data and data.get('numero_viaje'):
            fields_mapping['Número de Viaje'] = data['numero_viaje']
        
        if 'puertos' in data:
            fields_mapping['Puertos'] = ', '.join(data['puertos']) if isinstance(data['puertos'], list) else data['puertos']
        
        for field, value in fields_mapping.items():
            if value is not None:
                df_data['Campo'].append(field)
                df_data['Valor'].append(str(value))
        
        df = pd.DataFrame(df_data)
        
        # Crear archivo Excel con formato
        output_path_obj = Path(output_path)
        output_path_obj.parent.mkdir(parents=True, exist_ok=True)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Datos Extraídos', index=False)
            
            # Obtener la hoja para aplicar formato
            worksheet = writer.sheets['Datos Extraídos']
            
            # Ajustar ancho de columnas
            worksheet.column_dimensions['A'].width = 25
            worksheet.column_dimensions['B'].width = 40
            
            # Formato de encabezado
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        logger.info(f"Archivo Excel generado exitosamente: {output_path}")
        return output_path
    
    def export_multiple(self, data_list: List[Dict[str, any]], output_path: str, datos_generales: Dict = None) -> str:
        """
        Exporta múltiples registros a Excel.
        
        Args:
            data_list: Lista de diccionarios con datos normalizados (naves).
            output_path: Ruta donde guardar el archivo Excel.
            datos_generales: Datos generales del itinerario (opcional).
            
        Returns:
            Ruta del archivo generado.
        """
        if not data_list:
            raise ValueError("La lista de datos está vacía")
        
        output_path_obj = Path(output_path)
        output_path_obj.parent.mkdir(parents=True, exist_ok=True)
        
        # Preparar datos para DataFrame (campos requeridos siempre)
        rows = []
        for data in data_list:
            row = {
                'Naviera': data.get('naviera') or (datos_generales.get('naviera') if datos_generales else ''),
                'Nave': data.get('nombre') or data.get('nave_normalizada') or data.get('nombre_original'),
                'POL': data.get('pol') or (datos_generales.get('pol') if datos_generales else ''),
                'POD': data.get('pod') or (datos_generales.get('pod') if datos_generales else ''),
                'ETD': data.get('etd') or data.get('fecha_salida_normalizada') or (datos_generales.get('etd') if datos_generales else ''),
                'ETA': data.get('eta') or data.get('fecha_llegada_normalizada') or (datos_generales.get('eta') if datos_generales else ''),
                'Fecha': data.get('fecha_normalizada') or data.get('fecha_original'),
                'Semana': data.get('semana_normalizada') or (datos_generales.get('semana_normalizada') if datos_generales else ''),
                'Número de Viaje': data.get('numero_viaje', ''),
                'Número de Contenedor': data.get('numero_contenedor', ''),
                'Número de Booking': data.get('numero_booking', ''),
            }
            rows.append(row)
        
        df = pd.DataFrame(rows)
        
        # Crear archivo Excel
        output_path_obj = Path(output_path)
        output_path_obj.parent.mkdir(parents=True, exist_ok=True)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Itinerarios', index=False)
            
            # Aplicar formato
            worksheet = writer.sheets['Itinerarios']
            
            # Ajustar ancho de columnas
            for col in worksheet.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[col_letter].width = adjusted_width
            
            # Formato de encabezado
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        logger.info(f"Archivo Excel con {len(data_list)} registros generado: {output_path}")
        return output_path
